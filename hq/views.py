"""Vues HTTP : auth (user + pass + PIN), dashboard, API, board, calendrier, exports."""
from __future__ import annotations
import calendar as cal
import csv
import io
import json
from datetime import date, timedelta

from asgiref.sync import async_to_sync
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Avg, Count, Max, Sum
from django.http import (
    Http404, HttpResponse, HttpResponseRedirect, JsonResponse, StreamingHttpResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_GET, require_POST

from hq.geo import COUNTRY_CENTROIDS, country_name
from hq.job_manager import manager
from hq.models import (
    Comment, ContactMessage, Lead, ScrapeRun, Task, TaskLike, TaskVote, UserProfile,
)
from queries import list_categories, PLATFORM_GROUP_NAMES


LEAD_FIELDS = [
    "lead_score", "name", "role", "company",
    "emails", "phones", "email_candidates", "linkedin",
    "fund_size", "fund_close_step", "recency_months",
    "source", "source_url", "source_title", "evidence",
]


def _profile(user: User) -> UserProfile:
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile


# -----------------------------------------------------------------------
# Auth
# -----------------------------------------------------------------------
@csrf_protect
def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    error = None
    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        pin = (request.POST.get("pin") or "").strip()

        user = authenticate(request, username=username, password=password)
        if user is None:
            error = "Identifiants invalides."
        else:
            profile = _profile(user)
            if not profile.check_pin(pin):
                error = "Code PIN invalide."
            else:
                login(request, user)
                request.session.set_expiry(60 * 60 * 12)  # 12h
                nxt = request.GET.get("next") or (reverse("dashboard") + "?welcome=1")
                return HttpResponseRedirect(nxt)

    return render(request, "hq/login.html", {"error": error})


def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse("landing"))


# -----------------------------------------------------------------------
# Health
# -----------------------------------------------------------------------
def healthz(request):
    return HttpResponse("ok", content_type="text/plain")


# -----------------------------------------------------------------------
# Landing (public) + contact
# -----------------------------------------------------------------------
def landing_view(request):
    """Vitrine publique. Les utilisateurs connectés voient toujours la page ;
    un bouton les ramène à la console."""
    sent = request.GET.get("sent") == "1"
    return render(request, "hq/landing.html", {
        "sent": sent,
    })


@csrf_protect
@require_POST
def contact_submit(request):
    name = (request.POST.get("name") or "").strip()[:120]
    email = (request.POST.get("email") or "").strip()[:200]
    company = (request.POST.get("company") or "").strip()[:200]
    intent = (request.POST.get("intent") or "info").strip()
    message = (request.POST.get("message") or "").strip()[:5000]
    honeypot = (request.POST.get("website") or "").strip()

    if honeypot:  # bot caught
        return HttpResponseRedirect(reverse("landing") + "?sent=1#contact")

    valid_intents = {c[0] for c in ContactMessage.INTENT_CHOICES}
    if intent not in valid_intents:
        intent = "info"

    if not (name and email and message and "@" in email):
        return HttpResponseRedirect(reverse("landing") + "?err=1#contact")

    xff = request.META.get("HTTP_X_FORWARDED_FOR", "")
    ip = xff.split(",")[0].strip() if xff else request.META.get("REMOTE_ADDR")

    ContactMessage.objects.create(
        name=name, email=email, company=company,
        intent=intent, message=message, source_ip=ip or None,
    )
    return HttpResponseRedirect(reverse("landing") + "?sent=1#contact")


# -----------------------------------------------------------------------
# Dashboard (summary view) — `/dashboard/`
# -----------------------------------------------------------------------
@require_GET
@login_required
def dashboard(request):
    now = timezone.now()
    week_ago = now - timedelta(days=7)

    leads_qs = Lead.objects.all()
    runs_qs = ScrapeRun.objects.all()

    total_leads = leads_qs.count()
    leads_this_week = leads_qs.filter(created_at__gte=week_ago).count()
    total_runs = runs_qs.count()
    last_run = runs_qs.order_by("-started_at").first()

    avg_llm = leads_qs.exclude(llm_score__isnull=True).aggregate(v=Avg("llm_score"))["v"] or 0
    high_value = leads_qs.filter(llm_score__gte=80).count()

    countries_hit = (
        leads_qs.exclude(country="")
        .values("country")
        .annotate(n=Count("id"))
        .order_by("-n")[:8]
    )
    countries_hit = [
        {"iso2": c["country"], "name": country_name(c["country"]), "count": c["n"]}
        for c in countries_hit
    ]

    top_leads = (
        leads_qs.exclude(llm_score__isnull=True)
        .order_by("-llm_score", "-lead_score")[:6]
    )

    recent_runs = runs_qs.order_by("-started_at")[:5]

    # User's tasks / calendar preview
    upcoming_tasks = (
        Task.objects.filter(start_date__gte=now.date())
        .exclude(status="archived")
        .order_by("start_date")[:5]
    )
    my_open_tasks = (
        Task.objects.filter(status__in=["open", "in_progress"])
        .filter(author=request.user)
        .order_by("-updated_at")[:5]
    )

    # Access summary (what this user can use)
    access = {
        "scrape_categories": len(list_categories()),
        "platform_groups": len(PLATFORM_GROUP_NAMES),
        "llm_enabled": _llm_enabled(),
        "team_count": User.objects.count(),
    }

    return render(request, "hq/dashboard.html", {
        "now": now,
        "total_leads": total_leads,
        "leads_this_week": leads_this_week,
        "total_runs": total_runs,
        "last_run": last_run,
        "avg_llm": round(avg_llm or 0, 1),
        "high_value": high_value,
        "countries_hit": countries_hit,
        "top_leads": top_leads,
        "recent_runs": recent_runs,
        "upcoming_tasks": upcoming_tasks,
        "my_open_tasks": my_open_tasks,
        "access": access,
    })


# -----------------------------------------------------------------------
# Console (launcher + map + live log) — `/console/`
# -----------------------------------------------------------------------
@require_GET
@login_required
def console_view(request):
    recent = ScrapeRun.objects.all()[:10]
    return render(request, "hq/console.html", {
        "categories": list_categories(),
        "platform_groups": sorted(PLATFORM_GROUP_NAMES),
        "recent_runs": recent,
        "llm_enabled": _llm_enabled(),
    })


def _llm_enabled() -> bool:
    try:
        import config
        return bool(config.LLM_ENABLED)
    except Exception:
        return False


# -----------------------------------------------------------------------
# Profile
# -----------------------------------------------------------------------
@login_required
def profile_view(request):
    user = request.user
    profile = _profile(user)

    error = None
    notice = None
    if request.method == "POST":
        action = request.POST.get("action") or "info"

        if action == "info":
            display = (request.POST.get("display_name") or "").strip()[:80]
            color = (request.POST.get("color") or "").strip()[:16] or "#d4af6c"
            bio = (request.POST.get("bio") or "").strip()[:2000]
            profile.display_name = display
            profile.color = color
            profile.bio = bio
            profile.save()
            notice = "Profil mis à jour."

        elif action == "password":
            current = request.POST.get("current_password") or ""
            new1 = request.POST.get("new_password") or ""
            new2 = request.POST.get("new_password_confirm") or ""
            if not user.check_password(current):
                error = "Mot de passe actuel incorrect."
            elif not new1 or new1 != new2:
                error = "Les nouveaux mots de passe ne correspondent pas."
            elif len(new1) < 8:
                error = "Le mot de passe doit faire au moins 8 caractères."
            else:
                user.set_password(new1)
                user.save()
                login(request, user)  # keep session alive
                notice = "Mot de passe mis à jour."

        elif action == "pin":
            current_pin = (request.POST.get("current_pin") or "").strip()
            new_pin = (request.POST.get("new_pin") or "").strip()
            if not profile.check_pin(current_pin):
                error = "PIN actuel incorrect."
            elif not new_pin.isdigit() or len(new_pin) != 4:
                error = "Le nouveau PIN doit être composé de 4 chiffres."
            else:
                profile.set_pin(new_pin)
                profile.save()
                notice = "PIN mis à jour."

    authored = Task.objects.filter(author=user).count()
    assigned = Task.objects.filter(assignee=user).count()
    done = Task.objects.filter(assignee=user, status="done").count()
    return render(request, "hq/profile.html", {
        "profile": profile,
        "error": error,
        "notice": notice,
        "authored_count": authored,
        "assigned_count": assigned,
        "done_count": done,
    })


# -----------------------------------------------------------------------
# Board (wallpaper de tâches)
# -----------------------------------------------------------------------
@login_required
def board_view(request):
    status = request.GET.get("status") or ""
    sort = request.GET.get("sort") or "hot"
    qs = Task.objects.exclude(status="archived")
    if status in {"open", "in_progress", "done"}:
        qs = qs.filter(status=status)
    qs = qs.select_related("author", "assignee").prefetch_related("likes", "votes", "comments")

    tasks = list(qs)
    if sort == "new":
        tasks.sort(key=lambda t: t.created_at, reverse=True)
    elif sort == "votes":
        tasks.sort(key=lambda t: (t.score, t.created_at.timestamp()), reverse=True)
    else:  # hot = votes + likes + commentaires récents
        now_ts = timezone.now().timestamp()

        def heat(t):
            age_hours = max(1.0, (now_ts - t.created_at.timestamp()) / 3600.0)
            return (t.score * 2 + t.like_count + t.comments.count() * 0.5) / (age_hours ** 0.35)

        tasks.sort(key=heat, reverse=True)

    liked_ids = set(
        TaskLike.objects.filter(user=request.user, task__in=tasks).values_list("task_id", flat=True)
    )
    voted_map = {
        tv.task_id: tv.value
        for tv in TaskVote.objects.filter(user=request.user, task__in=tasks)
    }
    for t in tasks:
        t.user_liked = t.id in liked_ids
        t.user_vote = voted_map.get(t.id, 0)
    users = User.objects.order_by("username")

    return render(request, "hq/board.html", {
        "tasks": tasks,
        "users": users,
        "status": status,
        "sort": sort,
    })


@login_required
def task_create(request):
    if request.method != "POST":
        return redirect("board")
    title = (request.POST.get("title") or "").strip()[:200]
    description = (request.POST.get("description") or "").strip()
    assignee_id = request.POST.get("assignee") or ""
    start_raw = (request.POST.get("start_date") or "").strip()
    end_raw = (request.POST.get("end_date") or "").strip()
    if not title:
        return redirect("board")
    assignee = None
    if assignee_id:
        try:
            assignee = User.objects.get(id=int(assignee_id))
        except (User.DoesNotExist, ValueError):
            assignee = None
    start_d = _parse_date(start_raw)
    end_d = _parse_date(end_raw)
    if start_d and end_d and end_d < start_d:
        end_d = start_d
    t = Task.objects.create(
        title=title, description=description,
        author=request.user, assignee=assignee,
        start_date=start_d, end_date=end_d,
    )
    return redirect("task_detail", task_id=t.id)


@login_required
def task_detail(request, task_id: int):
    task = get_object_or_404(
        Task.objects.select_related("author", "assignee"),
        id=task_id,
    )
    if request.method == "POST":
        action = request.POST.get("action") or "comment"

        if action == "comment":
            body = (request.POST.get("body") or "").strip()
            if body:
                Comment.objects.create(task=task, author=request.user, body=body[:5000])
            return redirect("task_detail", task_id=task.id)

        if action == "status":
            new_status = request.POST.get("status") or "open"
            if new_status in {"open", "in_progress", "done", "archived"}:
                task.status = new_status
                task.save(update_fields=["status", "updated_at"])
            return redirect("task_detail", task_id=task.id)

        if action == "assign":
            aid = request.POST.get("assignee") or ""
            if aid:
                try:
                    task.assignee = User.objects.get(id=int(aid))
                except (User.DoesNotExist, ValueError):
                    task.assignee = None
            else:
                task.assignee = None
            task.save(update_fields=["assignee", "updated_at"])
            return redirect("task_detail", task_id=task.id)

        if action == "schedule":
            s = _parse_date(request.POST.get("start_date") or "")
            e = _parse_date(request.POST.get("end_date") or "")
            if s and e and e < s:
                e = s
            task.start_date = s
            task.end_date = e
            task.save(update_fields=["start_date", "end_date", "updated_at"])
            return redirect("task_detail", task_id=task.id)

        if action == "delete":
            if request.user == task.author:
                task.delete()
            return redirect("board")

    comments = task.comments.select_related("author").all()
    liked = TaskLike.objects.filter(task=task, user=request.user).exists()
    voted = TaskVote.objects.filter(task=task, user=request.user).first()
    users = User.objects.order_by("username")
    return render(request, "hq/task_detail.html", {
        "task": task,
        "comments": comments,
        "liked": liked,
        "voted_value": voted.value if voted else 0,
        "users": users,
    })


@login_required
@require_POST
def task_like(request, task_id: int):
    task = get_object_or_404(Task, id=task_id)
    like, created = TaskLike.objects.get_or_create(task=task, user=request.user)
    if not created:
        like.delete()
    return redirect(request.META.get("HTTP_REFERER") or reverse("board"))


@login_required
@require_POST
def task_vote(request, task_id: int):
    task = get_object_or_404(Task, id=task_id)
    try:
        value = int(request.POST.get("value") or "0")
    except ValueError:
        value = 0
    if value not in (-1, 1):
        return redirect(request.META.get("HTTP_REFERER") or reverse("board"))
    vote, created = TaskVote.objects.get_or_create(
        task=task, user=request.user, defaults={"value": value},
    )
    if not created:
        if vote.value == value:
            vote.delete()  # toggle off
        else:
            vote.value = value
            vote.save(update_fields=["value"])
    return redirect(request.META.get("HTTP_REFERER") or reverse("board"))


# -----------------------------------------------------------------------
# Calendrier (mois courant)
# -----------------------------------------------------------------------
@login_required
def calendar_view(request):
    today = timezone.localdate()
    try:
        year = int(request.GET.get("y") or today.year)
        month = int(request.GET.get("m") or today.month)
    except ValueError:
        year, month = today.year, today.month
    if month < 1 or month > 12:
        year, month = today.year, today.month

    first = date(year, month, 1)
    _, days_in_month = cal.monthrange(year, month)
    last = date(year, month, days_in_month)

    tasks = (
        Task.objects
        .filter(start_date__lte=last, end_date__gte=first)
        .select_related("author", "assignee")
    )
    # Fallback : tasks sans end_date (1 jour)
    tasks_single = (
        Task.objects
        .filter(start_date__gte=first, start_date__lte=last, end_date__isnull=True)
        .select_related("author", "assignee")
    )

    by_day: dict[date, list[Task]] = {}
    for t in list(tasks) + list(tasks_single):
        s = t.start_date
        e = t.end_date or t.start_date
        if not s:
            continue
        if s < first:
            s = first
        if e > last:
            e = last
        d = s
        while d <= e:
            by_day.setdefault(d, []).append(t)
            d = d.fromordinal(d.toordinal() + 1)

    # Grille semaines : Lundi → Dimanche
    cal_obj = cal.Calendar(firstweekday=0)
    weeks = []
    for week in cal_obj.monthdatescalendar(year, month):
        row = []
        for d in week:
            row.append({
                "date": d,
                "in_month": d.month == month,
                "tasks": by_day.get(d, []),
                "is_today": d == today,
            })
        weeks.append(row)

    prev_month = (first - timedelta(days=1)).replace(day=1)
    next_month = (last + timedelta(days=1)).replace(day=1)

    return render(request, "hq/calendar.html", {
        "year": year,
        "month": month,
        "month_label": first.strftime("%B %Y"),
        "weeks": weeks,
        "weekday_labels": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
        "prev_year": prev_month.year, "prev_month": prev_month.month,
        "next_year": next_month.year, "next_month": next_month.month,
        "today": today,
        "users": User.objects.order_by("username"),
    })


@login_required
@require_POST
def calendar_create(request):
    """Quick-add : title + date (+ optional end) depuis la vue calendrier."""
    title = (request.POST.get("title") or "").strip()[:200]
    start_raw = request.POST.get("start_date") or ""
    end_raw = request.POST.get("end_date") or ""
    if not title or not start_raw:
        return redirect("calendar")
    s = _parse_date(start_raw)
    e = _parse_date(end_raw) or s
    if not s:
        return redirect("calendar")
    if e and e < s:
        e = s
    Task.objects.create(
        title=title, author=request.user,
        start_date=s, end_date=e, status="open",
    )
    y = request.POST.get("return_y") or s.year
    m = request.POST.get("return_m") or s.month
    return redirect(f"{reverse('calendar')}?y={y}&m={m}")


def _parse_date(raw: str) -> date | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


# -----------------------------------------------------------------------
# API (scraper)
# -----------------------------------------------------------------------
@require_POST
@login_required
def api_start(request):
    try:
        payload = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "invalid json"}, status=400)

    categories = payload.get("categories") or None
    params = dict(
        categories=categories,
        min_priority=int(payload.get("min_priority", 5)),
        max_results_per_query=int(payload.get("max_results_per_query", 10)),
        use_llm=bool(payload.get("use_llm", True)),
        use_team_crawl=bool(payload.get("use_team_crawl", True)),
        use_email_enrich=bool(payload.get("use_email_enrich", True)),
        exclude_platforms=bool(payload.get("exclude_platforms", False)),
        platforms_only=bool(payload.get("platforms_only", False)),
        extra_geo=payload.get("geo") or None,
        pdf_only=bool(payload.get("pdf_only", False)),
        recency_months_max=int(payload.get("recency_months", 12)),
        recency_required=bool(payload.get("recency_required", False)),
        fetch_workers=int(payload.get("fetch_workers", 3)),
        extract_workers=int(payload.get("extract_workers", 2)),
        enrich_workers=int(payload.get("enrich_workers", 2)),
    )

    runner = async_to_sync(manager.start)(**params)
    return JsonResponse({"run_id": runner.id, "status": "running"})


@require_POST
@login_required
def api_stop(request):
    async_to_sync(manager.stop)()
    return JsonResponse({"status": "stopped"})


@require_GET
@login_required
def api_status(request):
    snap = manager.snapshot()
    snap["leads"] = manager.top_leads(limit=50)
    return JsonResponse(snap)


@require_GET
@login_required
def api_runs(request):
    qs = ScrapeRun.objects.all()[:30]
    data = [{
        "run_id": r.run_id,
        "started_at": r.started_at.isoformat(),
        "finished_at": r.finished_at.isoformat() if r.finished_at else None,
        "queries_total": r.queries_total,
        "people_unique": r.people_unique,
        "leads_final": r.leads_final,
        "status": r.status,
    } for r in qs]
    return JsonResponse({"runs": data})


@require_GET
@login_required
def api_leads_geo(request):
    """Return geo-located leads + per-country aggregates for the world map."""
    try:
        limit = max(1, min(int(request.GET.get("limit") or 600), 2000))
    except ValueError:
        limit = 600

    run_id = (request.GET.get("run_id") or "").strip()
    qs = Lead.objects.all()
    if run_id:
        qs = qs.filter(run__run_id=run_id)

    # Points (only leads we can place on the map)
    points_qs = (
        qs.exclude(lat__isnull=True)
        .exclude(lng__isnull=True)
        .order_by("-llm_score", "-lead_score")[:limit]
    )
    points = [{
        "id": l.id,
        "name": l.name,
        "role": l.role,
        "company": l.company,
        "country": l.country,
        "country_name": country_name(l.country),
        "city": l.city,
        "lat": l.lat,
        "lng": l.lng,
        "lead_score": round(l.lead_score or 0, 2),
        "llm_score": l.llm_score,
        "llm_score_reasoning": l.llm_score_reasoning,
        "company_description": l.company_description,
        "seniority": l.seniority,
        "fund_size": l.fund_size,
        "fund_close_step": l.fund_close_step,
        "source_url": l.source_url,
    } for l in points_qs]

    # Per-country aggregates
    agg = (
        qs.exclude(country="")
        .values("country")
        .annotate(
            count=Count("id"),
            avg_llm=Avg("llm_score"),
            top_llm=Max("llm_score"),
            avg_lead=Avg("lead_score"),
        )
        .order_by("-count")
    )
    countries = []
    for a in agg:
        iso2 = a["country"]
        centroid = COUNTRY_CENTROIDS.get(iso2.upper())
        countries.append({
            "iso2": iso2,
            "name": country_name(iso2),
            "count": a["count"],
            "avg_llm": round(a["avg_llm"] or 0, 1) if a["avg_llm"] is not None else None,
            "top_llm": a["top_llm"],
            "avg_lead": round(a["avg_lead"] or 0, 2) if a["avg_lead"] is not None else None,
            "lat": centroid[0] if centroid else None,
            "lng": centroid[1] if centroid else None,
        })

    return JsonResponse({
        "points": points,
        "countries": countries,
        "total": qs.count(),
        "placed": len(points),
    })


# -----------------------------------------------------------------------
# Exports : CSV + XLSX par run
# -----------------------------------------------------------------------
def _leads_queryset(run_id: str):
    run = get_object_or_404(ScrapeRun, run_id=run_id)
    qs = Lead.objects.filter(run=run).order_by("-lead_score")
    return run, qs


def _flatten(row: Lead) -> list:
    def j(v):
        if v is None:
            return ""
        if isinstance(v, list):
            return "; ".join(str(x) for x in v if x)
        return str(v)
    return [
        round(row.lead_score or 0, 3),
        row.name or "",
        row.role or "",
        row.company or "",
        j(row.emails),
        j(row.phones),
        j(row.email_candidates),
        row.linkedin or "",
        row.fund_size or "",
        row.fund_close_step or "",
        row.recency_months if row.recency_months is not None else "",
        row.source or "",
        row.source_url or "",
        row.source_title or "",
        (row.evidence or "").replace("\n", " ").replace("\r", " ")[:2000],
    ]


@require_GET
@login_required
def export_csv(request, run_id: str):
    run, qs = _leads_queryset(run_id)

    def stream():
        buf = io.StringIO()
        writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(LEAD_FIELDS)
        yield buf.getvalue(); buf.seek(0); buf.truncate()
        for lead in qs.iterator(chunk_size=500):
            writer.writerow(_flatten(lead))
            yield buf.getvalue(); buf.seek(0); buf.truncate()

    resp = StreamingHttpResponse(stream(), content_type="text/csv; charset=utf-8")
    fname = f"se-partners-hq_{run.run_id[:8]}_{run.started_at:%Y%m%d-%H%M}.csv"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@require_GET
@login_required
def export_xlsx(request, run_id: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    run, qs = _leads_queryset(run_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Leads {run.run_id[:8]}"

    header_fill = PatternFill("solid", fgColor="1A1407")
    header_font = Font(bold=True, color="D4AF6C", name="Calibri", size=11)
    ws.append(LEAD_FIELDS)
    for col_idx in range(1, len(LEAD_FIELDS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for lead in qs.iterator(chunk_size=500):
        ws.append(_flatten(lead))

    widths = [8, 24, 28, 28, 34, 22, 34, 40, 18, 18, 12, 10, 46, 46, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    meta = wb.create_sheet("Run")
    meta.append(["field", "value"])
    meta.append(["run_id", run.run_id])
    meta.append(["started_at", run.started_at.strftime("%Y-%m-%d %H:%M:%S UTC")])
    meta.append(["finished_at", run.finished_at.strftime("%Y-%m-%d %H:%M:%S UTC") if run.finished_at else ""])
    meta.append(["status", run.status])
    meta.append(["queries_total", run.queries_total])
    meta.append(["people_unique", run.people_unique])
    meta.append(["leads_final", run.leads_final])
    meta.append(["categories", ", ".join(run.categories or [])])
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fname = f"se-partners-hq_{run.run_id[:8]}_{run.started_at:%Y%m%d-%H%M}.xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp
